from product import Product
from openpyxl import load_workbook
import getpass

print("Do'kon bo'limi".center(135))
while True:
    print("\n===== MENU =====")
    print("1. Add Product")
    print("2. Sell Product")
    print("3. Report")
    print("0. Exit")

    menu = input("Tanlang: ")
######################################### Add Product ##############################################
    if menu == "1":
        family = input("Mahsulot oilasi: ")
        if family == '0':
            continue
        else:
            name = input("Mahsulot nomi: ")
            if name == '0':
                continue
            else:
                quantity = int(input("Miqdori: "))
                if quantity == '0':
                    continue
                mahsulot_bor = False

                wb = load_workbook("products.xlsx")

                if family in wb.sheetnames:
                    ws = wb[family]

                    for cell in ws["A"][1:]:
                        if cell.value == name:
                            mahsulot_bor = True
                            break
                wb.close()
                if mahsulot_bor:
                    row = cell.row
                    buy_price = ws[f"C{row}"].value
                    sell_price = ws[f"D{row}"].value
                    unit = ws[f"F{row}"].value
                else:
                    buy_price = int(input("Sotib olish narxi: "))
                    sell_price = int(input("Sotish narxi: "))
                    unit = input("Birligi: ")
                    product = Product(family,name,quantity,buy_price,sell_price,0,unit)
                    product.add_product()
######################################### Sell Product ##############################################
    elif menu == "2":
        family = input("Mahsulot oilasi: ")
        if family == '0':
            continue
        else:
            name = input("Mahsulot nomi: ")
            if name == '0':
                continue
            else:
                quantity = int(input("Nechta sotildi: "))
        product = Product(family,name,quantity,0,0,0,"")
        product.sell_product()
######################################### Report Database ##############################################
    elif menu == "3":
        login=input("Login: ")
        parol=getpass.getpass("Parol: ")
        if login == '1' and parol == '1':    
            while True:
                print("\n===== REPORT =====")
                print("1. Add Product Report")
                print("2. Sell Product Report")
                print("3. Database")
                print("0. Chiqish")
                menu_add=input("Tanlang: ")
                if menu_add == '1':
                    print(Product.get_add_history())
                elif menu_add == '2':
                    print(Product.get_sell_history())
                elif menu_add == '3':
                    print(Product.get_all_products())
                elif menu_add == '0':
                    break
######################################### Exit ##############################################
    elif menu == "0":
        break
    else:
        print("Noto'g'ri tanlov!")


