from main_abc import ProductStructure
from openpyxl import load_workbook

class Product(ProductStructure):
    def __init__(self, product_family, product_name, product_quantity, product_price, selling_price, product_sold, product_unit):
        self.product_family = product_family
        self.product_name = product_name
        self.product_quantity = product_quantity
        self.product_price = product_price
        self.selling_price = selling_price
        self.product_sold = product_sold
        self.product_unit = product_unit

    def get_product_family(self):
        return self.product_family
    
    def get_product_name(self):
        return self.product_name
    
    def get_product_quantity(self):
        return self.product_quantity

    def get_product_price(self):
        return self.product_price

    def get_selling_price(self):
        return self.selling_price

    def get_product_sold(self):
        return self.product_sold

    def get_product_unit(self):
        return self.product_unit
######################################### Add Product History ##############################################
    @staticmethod
    def save_add_history(product_name, product_quantity, product_price, product_unit):
        history_wb = load_workbook("history.xlsx")
        history_ws = history_wb["add_product"]

        history_ws.append([
            product_name,
            product_quantity,
            product_price,
            product_unit
        ])

        history_wb.save("history.xlsx")
        history_wb.close()
######################################### Add Product ##############################################
    def add_product(self):
        wb = load_workbook('products.xlsx')
        if self.product_family not in wb.sheetnames:
            wb.create_sheet(self.product_family)
            ws = wb[self.product_family]
            ws.append(["Name","Quantity" ,"Buy Price", "Sell Price", "Sold", "Unit"])
            ws.append([self.product_name, self.product_quantity, self.product_price, self.selling_price, self.product_sold, self.product_unit])
            wb.save('products.xlsx')
        elif self.product_family in wb.sheetnames:
            ws = wb[self.product_family]
            if self.product_name not in [cell.value for cell in ws['A'][1:]]:
                ws.append([self.product_name, self.product_quantity, self.product_price, self.selling_price, self.product_sold, self.product_unit])
                wb.save('products.xlsx')
            else:
                for cell in ws['A']:
                    if cell.value == self.product_name:
                        row = cell.row
                        current = int(ws[f'B{row}'].value)
                        ws[f'B{row}'] = current + self.product_quantity                        
                        print(f"{self.product_name} miqdori {current+self.product_quantity} ga o'zgartirildi va saqlandi.")
                        break
        wb.save('products.xlsx')  
        wb.close()

        Product.save_add_history(
            self.product_name,
            self.product_quantity,
            self.product_price,
            self.product_unit
        )
######################################### Add Product Report ##############################################
    @classmethod
    def get_add_history(cls):
        wb = load_workbook("history.xlsx", data_only=True)
        ws = wb["add_product"]

        qatorlar = []

        for row in ws.iter_rows(values_only=True):
            if any(row):
                qatorlar.append(list(row))

        if not qatorlar:
            wb.close()
            return []

        headers = qatorlar.pop(0)

        data_list = []
        for row in qatorlar:
            row_dict = dict(zip(headers, row))
            data_list.append(row_dict)

        wb.close()
        return data_list
######################################### Sell Product History ##############################################
    @staticmethod
    def save_sell_history(product_name, product_quantity, selling_price, product_unit):
        history_wb = load_workbook("history.xlsx")
        history_ws = history_wb["sell_product"]

        history_ws.append([
            product_name,
            product_quantity,
            selling_price,
            product_unit
        ])

        history_wb.save("history.xlsx")
        history_wb.close()   
######################################### Sell Product ##############################################
    def sell_product(self):
        wb = load_workbook('products.xlsx')
        if self.product_family not in wb.sheetnames:
            print(f"{self.product_family} kategoriyasi topilmadi.")
            wb.close()
            return
        ws = wb[self.product_family]
        row_num = None
        for cell in ws['A'][1:]:
            if cell.value == self.product_name:
                row_num = cell.row
                break
        if row_num is None:
            print(f"{self.product_name} topilmadi.")
            wb.close()
            return
        current_quantity = int(ws[f'B{row_num}'].value)
        if current_quantity < self.product_quantity:
            print(f"Omborda faqat {current_quantity} ta\n{self.product_name} mavjud.")
            wb.close()
            return
        ws[f'B{row_num}'] = current_quantity - self.product_quantity
        current_sold = int(ws[f'E{row_num}'].value)
        ws[f'E{row_num}'] = current_sold + self.product_quantity

        selling_price = ws[f'D{row_num}'].value
        product_unit = ws[f'F{row_num}'].value

        wb.save('products.xlsx')
        wb.close()
        print(f"{self.product_quantity} ta {self.product_name} sotildi.")

        Product.save_sell_history(
            self.product_name,
            self.product_quantity,
            selling_price,
            product_unit
        )
######################################### Sell Product Report ##############################################
    @classmethod
    def get_sell_history(cls):
        wb = load_workbook("history.xlsx", data_only=True)
        ws = wb["sell_product"]

        qatorlar = []

        for row in ws.iter_rows(values_only=True):
            if any(row):
                qatorlar.append(list(row))

        if not qatorlar:
            wb.close()
            return []

        headers = qatorlar.pop(0)

        data_list = []
        for row in qatorlar:
            row_dict = dict(zip(headers, row))
            data_list.append(row_dict)

        wb.close()
        return data_list
######################################### Database Report ##############################################
    @classmethod
    def get_all_products(cls):
        wb = load_workbook("products.xlsx", data_only=True)
        barcha_sahifalar = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sahifa_qatorlari = []
            for row in sheet.iter_rows(values_only=True):
                if any(row):
                    sahifa_qatorlari.append(list(row))
            if not sahifa_qatorlari:
                continue
            headers = sahifa_qatorlari.pop(0)
            data_list = []
            for row in sahifa_qatorlari:
                row_dict = dict(zip(headers, row))
                data_list.append(row_dict)
            barcha_sahifalar[sheet_name] = data_list
        wb.close()
        return barcha_sahifalar